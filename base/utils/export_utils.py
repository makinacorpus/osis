##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2016 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Color, Style, PatternFill
from django.utils.translation import ugettext_lazy as _

from base import models as mdl

HEADER = [str(_('academic_year')),
          str(_('session')),
          str(_('activity_code')),
          str(_('program')),
          str(_('registration_number')),
          str(_('lastname')),
          str(_('firstname')),
          str(_('numbered_score')),
          str(_('justification')),
          str(_('end_date')),
          str(_('ID'))]


def export_xls(academic_year_id, is_fac, exam_enrollments):
    academic_year = mdl.academic_year.find_academic_year_by_id(academic_year_id)

    wb = Workbook()
    ws = wb.active

    __columns_resizing(ws)
    ws.append(HEADER)

    enrollment_counter = 1
    for exam_enroll in exam_enrollments:
        student = exam_enroll.learning_unit_enrollment.student
        o = exam_enroll.learning_unit_enrollment.offer
        person = mdl.person.find_by_id(student.person.id)

        if exam_enroll.session_exam.offer_year_calendar.end_date is None:
            end_date = "-"
        else:
            end_date = exam_enroll.session_exam.offer_year_calendar.end_date.strftime('%d/%m/%Y')
        score = None
        if exam_enroll.score_final:
            if exam_enroll.session_exam.learning_unit_year.decimal_scores:
                score = "{0:.2f}".format(exam_enroll.score_final)
            else:
                score = "{0:.0f}".format(exam_enroll.score_final)
        justification = ""
        if exam_enroll.justification_final:
            justification = dict(mdl.exam_enrollment.JUSTIFICATION_TYPES)[exam_enroll.justification_final]
        ws.append([str(academic_year),
                   str(exam_enroll.session_exam.number_session),
                   exam_enroll.session_exam.learning_unit_year.acronym,
                   o.acronym,
                   student.registration_id,
                   person.last_name,
                   person.first_name,
                   score,
                   str(justification),
                   end_date,
                   exam_enroll.id])

        enrollment_counter += 1
        __coloring_non_editable(ws, enrollment_counter, score, exam_enroll.justification_final)

    number_session = list(exam_enrollments)[0].session_exam.number_session
    learn_unit_acronym = list(exam_enrollments)[0].session_exam.learning_unit_year.acronym

    ws.append([str(_('other_score_legend')),
               mdl.exam_enrollment.justification_label_authorized(is_fac)])
    filename = "session_%s_%s_%s.xls" % (str(academic_year.year),
                                         str(number_session),
                                         learn_unit_acronym)
    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    return response


def __columns_resizing(ws):
    """
    Definition of the columns sizes
    """
    col_academic_year = ws.column_dimensions['A']
    col_academic_year.width = 18
    col_academic_year = ws.column_dimensions['C']
    col_academic_year.width = 18
    col_academic_year = ws.column_dimensions['E']
    col_academic_year.width = 18
    col_last_name = ws.column_dimensions['F']
    col_last_name.width = 30
    col_first_name = ws.column_dimensions['G']
    col_first_name.width = 30
    col_note = ws.column_dimensions['H']
    col_note.width = 20
    col_note = ws.column_dimensions['I']
    col_note.width = 20
    col_id_exam_enrollment = ws.column_dimensions['K']
    # Hide the exam_enrollment_id column
    col_id_exam_enrollment.hidden = True


def __coloring_non_editable(ws, enrollment_counter, score, justification):
    """
    Coloring of the non-editable columns
    """
    style_no_modification = Style(fill=PatternFill(patternType='solid', fgColor=Color('C1C1C1')))
    column_number = 1
    while column_number < 12:
        if column_number < 8 or column_number > 9:
            ws.cell(row=enrollment_counter, column=column_number).style = style_no_modification
        else:
            if not(score is None and justification is None):
                ws.cell(row=enrollment_counter, column=8).style = style_no_modification
                ws.cell(row=enrollment_counter, column=9).style = style_no_modification

        column_number += 1
